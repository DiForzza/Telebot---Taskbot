import telebot
from telebot import types
import openpyxl
# @test4456_bot

bot = telebot.TeleBot('1530619842:AAH20gD_rs72BX9HsXlef2avhbxMPXUiROk')
wb_auth = openpyxl.load_workbook('sotr/auth.xlsx')
wb_task = openpyxl.load_workbook('sotr/task.xlsx')
sheet_auth = wb_auth.active
sheet_task = wb_task.active
rows_auth = sheet_auth.max_row
rows_task = sheet_task.max_row
cols = sheet_auth.max_column
auth_ok = 0

def auth(phone_usm, chatid):
    global auth_ok
    for i in range(2, rows_auth + 1):
        surname = sheet_auth.cell(row=i, column=1)
        name = sheet_auth.cell(row=i, column=2)
        basenumber = sheet_auth.cell(row=i, column=3)
        if int(phone_usm) == basenumber.value:
            auth_ok = 1
            bot.send_message(chatid, f'Доброго времени суток, {surname.value} {name.value}.')
            c1 = sheet_auth.cell(row=i, column=5)
            c1.value = chatid
            wb_auth.save("sotr/auth.xlsx")
    return auth_ok

@bot.message_handler(commands=["start"])
def geophone(message):
    is_auth_ok(message)
    if is_auth_ok == False:
        keyboard1 = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        button_phone = types.KeyboardButton(text="Отправить номер телефона", request_contact=True)
        keyboard1.add(button_phone)
        bot.send_message(message.chat.id, "Пройдите авторизацию по номеру телефона", reply_markup=keyboard1)
 #   else:
 #       reply_markup = keyboard2

def is_auth_ok(message):
    chatid = message.chat.id
    for i in range(2, rows_auth + 1):
        if sheet_auth.cell(row=i, column=5) is not None:
            bot.send_message(chatid, f'Авторизация уже выполнена.')
            keyboard2 = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
            button_task = types.KeyboardButton(text="Просмотр своих задач")
            button_settaskto = types.KeyboardButton(text="Поставить задачу")
            button_mytaskto = types.KeyboardButton(text="Мои назначенные задачи")
            button_settask = types.KeyboardButton(text="Отметить задачу выполненной")
            keyboard2.add(button_task, button_settaskto)
            keyboard2.add(button_mytaskto, button_settask)
            bot.send_message(message.chat.id, "Выберите интересующий пункт меню:", reply_markup=keyboard2)
            return True

@bot.message_handler(content_types=['contact'])
def read_contact_phone(message):
    global auth_ok
    auth_ok = 0
    phone_usm = message.contact.phone_number
    chatid = message.chat.id
    auth(phone_usm, chatid)
    if auth_ok == 0:
        bot.send_message(chatid, f'Авторизация не пройдена. Номера не существует. {auth_ok}')
    else:
        keyboard2 = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        button_task = types.KeyboardButton(text="Просмотр своих задач")
        button_settaskto = types.KeyboardButton(text="Поставить задачу")
        button_mytaskto = types.KeyboardButton(text="Мои назначенные задачи")
        button_settask = types.KeyboardButton(text="Отметить задачу выполненной")
        keyboard2.add(button_task, button_settaskto)
        keyboard2.add(button_mytaskto, button_settask)
        bot.send_message(message.chat.id, "Выберите интересующий пункт меню:", reply_markup=keyboard2)

def settask_find_coincidence(message):
    chat_id = message.chat.id
    surname = message.text
    surname_founded = 0
    if message.text != 'Поставить задачу':
        for number_in_base in range(1, rows_auth + 1):
            surnameinbase = sheet_auth.cell(row=number_in_base, column=1)
            if surname == surnameinbase.value:
                if sheet_auth.cell(row=number_in_base, column=5).value is not None:
                    msg = bot.send_message(chat_id, f'В базе найден {surname}. Напишите задание: ')
                    bot.register_next_step_handler(msg, settask_write_to_base, number_in_base)
                    surname_founded = 1
                    break
                else:
                    surname_founded = 2
                    break
        if surname_founded == 0:
            bot.send_message(chat_id, f'В базе фамилия не найдена.')
        elif surname_founded == 2:
            bot.send_message(chat_id, f'Пользователь не зарегистрирован')
    else:
        seemytask(message)
        return

def find_my_surname(chat_id):
    for i in range(1, rows_auth + 1):
        id = sheet_auth.cell(row=i, column=5).value
        if id == chat_id:
            my_surname = sheet_auth.cell(row=i, column=1).value
            #bot.send_message(chat_id, my_surname)
            return my_surname

def settask_write_to_base(message, number_in_base):
    chat_id = message.chat.id
    task = message.text
    number_of_cols = 0
    #bot.send_message(chat_id, sheet_task.max_column)
    for i in range(1, sheet_task.max_column + 1):
        comparison = sheet_task.cell(row=number_in_base, column=i)
        if comparison.value is not None:
            number_of_cols += 1
    bot.send_message(chat_id, f'Вы написали задание: {task}, номер строки {number_in_base}, теперь {number_of_cols + 1} заданий ')
    c1 = sheet_task.cell(row=number_in_base, column=number_of_cols +1)
    c1.value = task
    wb_task.save("sotr/task.xlsx")
    send_task_to_id = sheet_auth.cell(row=number_in_base, column=5).value
    bot.send_message(send_task_to_id, f'Вы получили новое задание от {find_my_surname(chat_id)}: ')
    #bot.send_message(send_task_to_id, f'{message} {task}')


def seemytask(message):
    chat_id = message.chat.id
    bot.send_message(chat_id, f'Привет, {message.chat.id}, {message.from_user.id}')  # 814835614 Dima

@bot.message_handler(content_types=['text'])
def send_text(message):
    if message.text == 'Просмотр своих задач':
        seemytask(message)
    elif message.text == 'Поставить задачу':
        msg =  bot.send_message(message.chat.id, 'Напишите фамилию того, для кого Вы создаете задачу:')
        bot.register_next_step_handler(msg, settask_find_coincidence)
    elif message.text == 'Мои назначенные задачи':
        bot.send_message(message.chat.id, 'Мои задачи')
    elif message.text == 'Отметить задачу выполненной':
        bot.send_message(message.chat.id, 'Отметить задачу')

bot.polling()
